from openpyxl import load_workbook
wb = load_workbook('Trunk Groups.xlsx')

sheet = wb['Sheet1']

search_TG = 'TATXN'

found = False
for row in sheet.iter_rows():
    for cell in row:
        if cell.value == search_TG:
            operator = sheet.cell(row=cell.row, column=1).value  # This is A1
            print(operator)

            found = True
            break  # Stop after finding the first match
    if found:
        break

if not found:
    print(f"Value '{search_TG}' not found in the sheet.")